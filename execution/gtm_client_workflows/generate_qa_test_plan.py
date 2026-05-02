#!/usr/bin/env python3
"""
generate_qa_test_plan.py
description: Generates a professional XLSX QA test plan for the Accessory Masters
             cold email system. 5 sheets: Executive Summary, Test Cases, Traceability
             Matrix, Blocked Items, Execution Log.
inputs: --output (default .tmp/accessory_masters_qa_test_plan.xlsx)
outputs: XLSX file at the specified path
usage:
    py execution/gtm_client_workflows/generate_qa_test_plan.py
    py execution/gtm_client_workflows/generate_qa_test_plan.py --output .tmp/custom_name.xlsx
"""

import argparse
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent.parent


# ---------------------------------------------------------------------------
# Test case data
# ---------------------------------------------------------------------------

def build_all_test_cases() -> list[dict]:
    """Return every test case as a list of dicts."""
    return [
        # ── UNIT TESTS ────────────────────────────────────────────────
        # Reply Classifier
        {"id": "TC-UNIT-001", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: hot_positive on phone number", "desc": "Classify 'Call me at 555-1234' as hot_positive", "pre": "None", "steps": "Call classify_mock('Call me at 555-1234')", "expected": "Returns 'hot_positive'", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-002", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: hot_positive on ready to sell", "desc": "Classify 'I'm ready to sell my business' as hot_positive", "pre": "None", "steps": "Call classify_mock('I\\'m ready to sell my business')", "expected": "Returns 'hot_positive'", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-003", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: hot_positive on schedule call", "desc": "Classify 'Let's schedule a call' as hot_positive", "pre": "None", "steps": "Call classify_mock('Let\\'s schedule a call to discuss')", "expected": "Returns 'hot_positive'", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-004", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: negative on not interested", "desc": "Classify 'Not interested, thanks' as negative", "pre": "None", "steps": "Call classify_mock('Not interested, thanks')", "expected": "Returns 'negative'", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-005", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: negative on unsubscribe", "desc": "Classify 'Unsubscribe me' as negative", "pre": "None", "steps": "Call classify_mock('Unsubscribe me immediately')", "expected": "Returns 'negative'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-006", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: neutral on out of office", "desc": "Classify out-of-office as neutral", "pre": "None", "steps": "Call classify_mock('I am out of office until next week')", "expected": "Returns 'neutral'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-007", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: positive on interested", "desc": "Classify 'interested in learning more' as positive", "pre": "None", "steps": "Call classify_mock('I\\'m interested in learning more')", "expected": "Returns 'positive'", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-008", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: unknown text defaults neutral", "desc": "Gibberish input defaults to neutral", "pre": "None", "steps": "Call classify_mock('kjashdfkjhsdf')", "expected": "Returns 'neutral'", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-009", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: case insensitive matching", "desc": "Classification works regardless of case", "pre": "None", "steps": "Call classify_mock('NOT INTERESTED')", "expected": "Returns 'negative'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-010", "cat": "Unit Test", "deliv": "D6", "name": "classify_mock: custom signal overrides", "desc": "Custom signal dict overrides defaults", "pre": "None", "steps": "Call classify_mock with custom signals dict", "expected": "Uses custom signals for matching", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        # Auto-Reply
        {"id": "TC-UNIT-011", "cat": "Unit Test", "deliv": "D6", "name": "handle_reply: handoff for hot lead", "desc": "Hot positive reply triggers human handoff", "pre": "Config with auto_reply enabled", "steps": "Call handle_reply with hot_positive reply, mock=True", "expected": "action=handoff, reason contains 'hot lead'", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-012", "cat": "Unit Test", "deliv": "D6", "name": "handle_reply: auto_reply for positive", "desc": "Positive reply gets auto-reply with text and delay", "pre": "Config with auto_reply enabled", "steps": "Call handle_reply with positive reply, mock=True", "expected": "action=auto_reply, reply_text non-empty, delay 120-420s", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-013", "cat": "Unit Test", "deliv": "D6", "name": "handle_reply: skip when disabled", "desc": "Disabled auto_reply config skips all processing", "pre": "Config with auto_reply.enabled=false", "steps": "Call handle_reply with any reply", "expected": "action=skip, reason contains 'disabled'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-014", "cat": "Unit Test", "deliv": "D6", "name": "handle_reply: skip for negative", "desc": "Negative classification skips auto-reply", "pre": "Config with auto_reply enabled", "steps": "Call handle_reply with negative reply", "expected": "action=skip, reason contains 'not actionable'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-015", "cat": "Unit Test", "deliv": "D6", "name": "auto_reply delay within 120-420s range", "desc": "Delay is randomized within configured range", "pre": "Config with delay_min=120, delay_max=420", "steps": "Call handle_reply, check delay_seconds", "expected": "120 <= delay_seconds <= 420", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-016", "cat": "Unit Test", "deliv": "D6", "name": "should_handoff: detects phone number", "desc": "Phone number in body triggers handoff", "pre": "None", "steps": "Call should_handoff('Call me at 555-1234')", "expected": "Returns True", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-017", "cat": "Unit Test", "deliv": "D6", "name": "should_handoff: custom signals list", "desc": "Custom hot lead signals override defaults", "pre": "None", "steps": "Call should_handoff with custom_signals=['magic word']", "expected": "Matches custom signal, returns True", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        # Telegram
        {"id": "TC-UNIT-018", "cat": "Unit Test", "deliv": "D8", "name": "telegram: format includes lead fields", "desc": "Formatted message contains name, email, company", "pre": "Sample reply dict", "steps": "Call format_positive_reply(reply)", "expected": "Message contains from_name, from_email, company", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-019", "cat": "Unit Test", "deliv": "D8", "name": "telegram: body truncated to 200 chars", "desc": "Long reply bodies are truncated in notification", "pre": "Reply with 500-char body", "steps": "Call format_positive_reply with long body", "expected": "Body preview limited to 200 characters", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-020", "cat": "Unit Test", "deliv": "D8", "name": "telegram: handles missing fields gracefully", "desc": "Empty dict input produces valid message with defaults", "pre": "None", "steps": "Call format_positive_reply({})", "expected": "Returns string with 'Unknown' defaults", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        # Report Generator
        {"id": "TC-UNIT-021", "cat": "Unit Test", "deliv": "D9", "name": "report: mock metrics have required keys", "desc": "generate_mock_metrics returns instantly + ghl data", "pre": "None", "steps": "Call generate_mock_metrics()", "expected": "Dict with 'instantly' and 'ghl' keys, all subfields present", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-022", "cat": "Unit Test", "deliv": "D9", "name": "report: generate_report combines metrics", "desc": "Report structure contains email + crm + summary", "pre": "Mock metrics", "steps": "Call generate_report(instantly, ghl)", "expected": "Dict with client, date_range, email, crm, summary", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-023", "cat": "Unit Test", "deliv": "D9", "name": "report: HTML contains metric values", "desc": "HTML report embeds actual numbers", "pre": "Metrics with known values", "steps": "Call format_html_report(report), check for value strings", "expected": "HTML contains emails_sent, deliverability%, pipeline value", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-UNIT-024", "cat": "Unit Test", "deliv": "D9", "name": "report: Telegram format contains metrics", "desc": "Telegram report includes key numbers", "pre": "Metrics with known values", "steps": "Call format_telegram_report(report)", "expected": "Message contains emails_sent, replies, pipeline value", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        # Variant Generator
        {"id": "TC-UNIT-025", "cat": "Unit Test", "deliv": "D12", "name": "variant: validate rejects exclamation marks", "desc": "Variant with ! fails validation", "pre": "Constraints with no_exclamation_marks=True", "steps": "Call validate_variant('Great opportunity!', constraints)", "expected": "Returns False", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_variant_generator.py", "notes": ""},

        # ── INTEGRATION TESTS ─────────────────────────────────────────
        {"id": "TC-INTG-001", "cat": "Integration", "deliv": "D-ALL", "name": "Module import: pipeline_utils", "desc": "pipeline_utils imports without errors", "pre": "Python environment configured", "steps": "import modules.pipeline_utils", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-002", "cat": "Integration", "deliv": "D6", "name": "Module import: reply_classifier", "desc": "reply_classifier imports without errors", "pre": "Python environment configured", "steps": "import modules.reply_classifier", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-003", "cat": "Integration", "deliv": "D3", "name": "Module import: llm_client", "desc": "llm_client imports without errors", "pre": "Python environment configured", "steps": "import modules.llm_client", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-004", "cat": "Integration", "deliv": "D11", "name": "Module import: instantly", "desc": "instantly module imports without errors", "pre": "Python environment configured", "steps": "import modules.outputs.instantly", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-005", "cat": "Integration", "deliv": "D5", "name": "Module import: ghl", "desc": "GHL module imports without errors", "pre": "Python environment configured", "steps": "import modules.outputs.ghl", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-006", "cat": "Integration", "deliv": "D8", "name": "Module import: telegram", "desc": "telegram module imports without errors", "pre": "Python environment configured", "steps": "import modules.outputs.telegram", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-007", "cat": "Integration", "deliv": "D6", "name": "Module import: auto_reply", "desc": "auto_reply module imports without errors", "pre": "Python environment configured", "steps": "import modules.outputs.auto_reply", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-008", "cat": "Integration", "deliv": "D9", "name": "Module import: report_generator", "desc": "report_generator module imports without errors", "pre": "Python environment configured", "steps": "import modules.outputs.report_generator", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-009", "cat": "Integration", "deliv": "D1", "name": "Module import: serper_maps_scraper", "desc": "Serper module imports without errors", "pre": "Python environment configured", "steps": "import lead_sourcing.serper_maps_scraper", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-010", "cat": "Integration", "deliv": "D1", "name": "Module import: prospeo_leads", "desc": "Prospeo module imports without errors", "pre": "Python environment configured", "steps": "import lead_sourcing.prospeo_leads", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-011", "cat": "Integration", "deliv": "D2", "name": "Module import: anymailfinder_lookup", "desc": "AnymailFinder module imports without errors", "pre": "Python environment configured", "steps": "import enrichment.anymailfinder_lookup", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-012", "cat": "Integration", "deliv": "D2", "name": "Module import: million_verifier", "desc": "Million Verifier module imports without errors", "pre": "Python environment configured", "steps": "import enrichment.million_verifier", "expected": "No ImportError", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-013", "cat": "Integration", "deliv": "D-ALL", "name": "Config: required top-level keys present", "desc": "accessory_masters.json has icp, sourcing, enrichment, etc.", "pre": "Config file exists", "steps": "Load config, check for 9 required keys", "expected": "All keys present: icp, sourcing, enrichment, personalization, instantly, ghl, classification, auto_reply, reporting", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-014", "cat": "Integration", "deliv": "D-ALL", "name": "Config: tone.json has voice and never_say", "desc": "Tone config contains required keys", "pre": "tone.json exists", "steps": "Load tone.json, check for voice, never_say, example_openers", "expected": "All keys present", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},
        {"id": "TC-INTG-015", "cat": "Integration", "deliv": "D-ALL", "name": "pipeline_utils: all functions callable", "desc": "Exported functions exist and are callable", "pre": "Module imported", "steps": "Check callable() for retry_with_backoff, deduplicate, load_config, save_leads, load_leads, normalize_domain, compute_dedup_key", "expected": "All 7 functions callable", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_integration.py", "notes": ""},

        # ── E2E PIPELINE TESTS ────────────────────────────────────────
        {"id": "TC-E2E-001", "cat": "E2E Pipeline", "deliv": "D11", "name": "Full pipeline mock run completes", "desc": "run_pipeline with stage=all, mock=True, force=True succeeds", "pre": "Config loaded, tmp_path available", "steps": "Call run_pipeline(config, stage='all', mock=True, force=True)", "expected": "No exceptions raised", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-002", "cat": "E2E Pipeline", "deliv": "D11", "name": "Pipeline state file created after run", "desc": "pipeline_state.json exists after pipeline completes", "pre": "Pipeline run complete", "steps": "Check tmp_path / 'pipeline_state.json' exists", "expected": "File exists", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-003", "cat": "E2E Pipeline", "deliv": "D11", "name": "Pipeline state file is valid JSON", "desc": "State file parses as valid JSON dict", "pre": "State file created", "steps": "json.load(state_file)", "expected": "Returns dict", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-004", "cat": "E2E Pipeline", "deliv": "D11", "name": "poll_replies mock completes without error", "desc": "Reply polling in mock mode processes all reply types", "pre": "Config loaded", "steps": "Call poll_replies(config, mock=True)", "expected": "No exceptions raised", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-005", "cat": "E2E Pipeline", "deliv": "D1", "name": "Serper get_mock_data returns list", "desc": "Mock data function returns non-empty list", "pre": "Module imported", "steps": "Call get_mock_data()", "expected": "Returns list with len > 0", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-006", "cat": "E2E Pipeline", "deliv": "D1", "name": "Serper mock leads have required fields", "desc": "Each mock lead has business_name, address, phone, domain", "pre": "Mock data loaded", "steps": "Check required fields on each lead", "expected": "No missing fields", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-007", "cat": "E2E Pipeline", "deliv": "D1", "name": "parse_serper_results transforms raw places", "desc": "Raw API places converted to lead dicts", "pre": "Raw place data with title, address, phone", "steps": "Call parse_serper_results with synthetic raw data", "expected": "Returns list of leads with business_name field", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-008", "cat": "E2E Pipeline", "deliv": "D1", "name": "Prospeo mock data returns list with entries", "desc": "Mock data returns non-empty list", "pre": "Module imported", "steps": "Call get_mock_data()", "expected": "Returns list with len > 0, entries have business_name or first_name", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-009", "cat": "E2E Pipeline", "deliv": "D1", "name": "parse_prospeo_results with V2 data", "desc": "V2 nested person/company structure parsed correctly", "pre": "Synthetic V2 data", "steps": "Call parse_prospeo_results with nested person/company dict", "expected": "Returns lead with owner_email field populated", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-010", "cat": "E2E Pipeline", "deliv": "D3", "name": "Mock opener returns nonempty string", "desc": "get_mock_opener produces valid text", "pre": "Lead dict provided", "steps": "Call get_mock_opener(lead)", "expected": "Returns non-empty string", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_e2e.py", "notes": ""},
        {"id": "TC-E2E-011", "cat": "E2E Pipeline", "deliv": "D11", "name": "Mock replies have required fields", "desc": "All mock replies contain from_email, from_name, subject, body, company, industry, received_at", "pre": "Pipeline module imported", "steps": "Call _get_mock_replies(), check required fields", "expected": "No missing fields in any reply", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},
        {"id": "TC-E2E-012", "cat": "E2E Pipeline", "deliv": "D11", "name": "Mock replies classified into all 4 categories", "desc": "Mock replies produce hot_positive, positive, negative, and neutral", "pre": "Mock replies + classifier loaded", "steps": "Classify all mock replies, collect unique classifications", "expected": "Set contains all 4 categories", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_pipeline.py", "notes": ""},

        # ── SANITY / SMOKE TESTS ──────────────────────────────────────
        {"id": "TC-SANE-001", "cat": "Sanity", "deliv": "D-ALL", "name": ".env file exists", "desc": "Environment file present in project root", "pre": "None", "steps": "Check (ROOT / '.env').exists()", "expected": "True", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-002", "cat": "Sanity", "deliv": "D-ALL", "name": "accessory_masters.json exists", "desc": "Main config file present", "pre": "None", "steps": "Check config/accessory_masters.json exists", "expected": "True", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-003", "cat": "Sanity", "deliv": "D-ALL", "name": "tone.json exists", "desc": "Tone config file present", "pre": "None", "steps": "Check config/tone.json exists", "expected": "True", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-004", "cat": "Sanity", "deliv": "D4", "name": "_redirects file exists", "desc": "CF Pages redirect file present", "pre": "None", "steps": "Check website/_redirects exists", "expected": "True", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-005", "cat": "Sanity", "deliv": "D10", "name": "wrangler.toml exists", "desc": "CF Worker config present", "pre": "None", "steps": "Check execution/infrastructure/api-proxy/wrangler.toml exists", "expected": "True", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-006", "cat": "Sanity", "deliv": "D-ALL", "name": ".gitignore exists", "desc": "Gitignore file present", "pre": "None", "steps": "Check .gitignore exists", "expected": "True", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-007", "cat": "Sanity", "deliv": "D-ALL", "name": "All 10 execution scripts exist", "desc": "Every registered script file present on disk", "pre": "None", "steps": "Check 10 parameterized script paths", "expected": "All 10 files exist", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-008", "cat": "Sanity", "deliv": "D-ALL", "name": "Config files are valid JSON", "desc": "All JSON configs parse without error", "pre": "None", "steps": "json.load() on accessory_masters.json, tone.json, vercel.json", "expected": "No JSON decode errors", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-009", "cat": "Sanity", "deliv": "D10", "name": "_redirects has worker URL", "desc": "Redirect file points to correct CF Worker", "pre": "_redirects exists", "steps": "Read file, check for worker URL string", "expected": "Contains 'accessory-masters-api.accessory-masters.workers.dev'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-010", "cat": "Sanity", "deliv": "D10", "name": "wrangler.toml has account ID", "desc": "Worker config has correct CF account", "pre": "wrangler.toml exists", "steps": "Read file, check for account_id", "expected": "Contains '26e5b8612be35e5d23a9186fcf5288d0'", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-011", "cat": "Sanity", "deliv": "D-ALL", "name": ".env is gitignored", "desc": "Secret file excluded from version control", "pre": ".gitignore exists", "steps": "Check .gitignore contains '.env'", "expected": "Found", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-012", "cat": "Sanity", "deliv": "D-ALL", "name": "credentials.json is gitignored", "desc": "Google creds excluded from version control", "pre": ".gitignore exists", "steps": "Check .gitignore contains 'credentials.json'", "expected": "Found", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-013", "cat": "Sanity", "deliv": "D-ALL", "name": ".tmp/ is gitignored", "desc": "Temp files excluded from version control", "pre": ".gitignore exists", "steps": "Check .gitignore contains '.tmp/'", "expected": "Found", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-014", "cat": "Sanity", "deliv": "D-ALL", "name": "No hardcoded API keys in Python source", "desc": "No key patterns found in execution/*.py", "pre": "Source files exist", "steps": "Scan all .py files for 4 forbidden key-prefix patterns (see test_sanity.py FORBIDDEN_PATTERNS)", "expected": "Zero matches", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SANE-015", "cat": "Sanity", "deliv": "D-ALL", "name": "PRD directive file exists", "desc": "Master GTM directive present", "pre": "None", "steps": "Check directives/gtm_client_workflows/accessory_masters_gtm.md exists", "expected": "True", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},

        # ── EDGE CASE / MONKEY TESTS ──────────────────────────────────
        {"id": "TC-EDGE-001", "cat": "Edge Case", "deliv": "D6", "name": "classify_mock: None input returns neutral", "desc": "None body defaults to neutral", "pre": "None", "steps": "Call classify_mock(None)", "expected": "Returns 'neutral'", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-002", "cat": "Edge Case", "deliv": "D6", "name": "classify_mock: empty string returns neutral", "desc": "Empty body defaults to neutral", "pre": "None", "steps": "Call classify_mock('')", "expected": "Returns 'neutral'", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-003", "cat": "Edge Case", "deliv": "D6", "name": "classify_mock: integer input raises AttributeError", "desc": "Non-string input raises expected error", "pre": "None", "steps": "Call classify_mock(12345)", "expected": "Raises AttributeError", "pri": "P3", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-004", "cat": "Edge Case", "deliv": "D6", "name": "classify_mock: very long string handled", "desc": "5000-word input doesn't crash", "pre": "None", "steps": "Call classify_mock('interested ' * 5000)", "expected": "Returns valid classification", "pri": "P3", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-005", "cat": "Edge Case", "deliv": "D6", "name": "classify_mock: unicode input handled", "desc": "Unicode text doesn't crash", "pre": "None", "steps": "Call classify_mock('Je suis interesse')", "expected": "Returns valid classification", "pri": "P3", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-006", "cat": "Edge Case", "deliv": "D6", "name": "classify_mock: special characters return neutral", "desc": "Symbols-only input defaults to neutral", "pre": "None", "steps": "Call classify_mock('!@#$%^&*()')", "expected": "Returns 'neutral'", "pri": "P3", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-007", "cat": "Edge Case", "deliv": "D6", "name": "should_handoff: None and empty return False", "desc": "Missing body doesn't trigger handoff", "pre": "None", "steps": "Call should_handoff(None), should_handoff('')", "expected": "Both return False", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-008", "cat": "Edge Case", "deliv": "D8", "name": "format_positive_reply: empty dict", "desc": "Empty reply dict produces valid output", "pre": "None", "steps": "Call format_positive_reply({})", "expected": "Returns non-empty string", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-009", "cat": "Edge Case", "deliv": "D8", "name": "format_positive_reply: None values", "desc": "Dict with None values produces valid output", "pre": "None", "steps": "Call format_positive_reply with None values", "expected": "Returns string, no crash", "pri": "P3", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-010", "cat": "Edge Case", "deliv": "D6", "name": "handle_reply: disabled config skips", "desc": "auto_reply disabled returns skip action", "pre": "Config with enabled=False", "steps": "Call handle_reply with disabled config", "expected": "action=skip", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-011", "cat": "Edge Case", "deliv": "D6", "name": "generate_reply_mock: empty body", "desc": "Empty body still produces valid reply", "pre": "None", "steps": "Call generate_reply_mock('', 'context')", "expected": "Returns non-empty string", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-012", "cat": "Edge Case", "deliv": "D6", "name": "generate_reply_mock: very long body", "desc": "1000x repeated text doesn't crash", "pre": "None", "steps": "Call generate_reply_mock('Tell me about selling ' * 1000, 'context')", "expected": "Returns non-empty string", "pri": "P3", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-013", "cat": "Edge Case", "deliv": "D1", "name": "parse_serper_results: empty places list", "desc": "Empty input returns empty list", "pre": "None", "steps": "Call parse_serper_results([], ...)", "expected": "Returns []", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-014", "cat": "Edge Case", "deliv": "D1", "name": "parse_serper_results: places with missing keys", "desc": "Incomplete place data still produces lead", "pre": "None", "steps": "Call parse_serper_results([{'title': 'Test'}], ...)", "expected": "Returns list with 1 lead", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-015", "cat": "Edge Case", "deliv": "D1", "name": "parse_prospeo_results: empty results", "desc": "Empty input returns empty list", "pre": "None", "steps": "Call parse_prospeo_results([], ...)", "expected": "Returns []", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-016", "cat": "Edge Case", "deliv": "D1", "name": "parse_prospeo_results: partial data", "desc": "Incomplete records still produce leads", "pre": "None", "steps": "Call parse_prospeo_results with partial dicts", "expected": "Returns list", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-017", "cat": "Edge Case", "deliv": "D-ALL", "name": "deduplicate: empty list returns empty", "desc": "Empty input handled gracefully", "pre": "None", "steps": "Call deduplicate([])", "expected": "Returns []", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},
        {"id": "TC-EDGE-018", "cat": "Edge Case", "deliv": "D-ALL", "name": "deduplicate: all duplicates return 1", "desc": "Identical records collapsed to single entry", "pre": "None", "steps": "Call deduplicate([lead, lead.copy(), lead.copy()])", "expected": "Returns list with 1 item", "pri": "P2", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_monkey.py", "notes": ""},

        # ── MANUAL QA / UAT TESTS ─────────────────────────────────────
        {"id": "TC-UAT-001", "cat": "UAT", "deliv": "D4", "name": "Website loads at production URL", "desc": "All 3 pages load without errors", "pre": "CF Pages deployed", "steps": "1. Navigate to https://accessory-masters-site.pages.dev\n2. Click through index, dashboard, signup pages\n3. Verify no broken images or layout issues", "expected": "All 3 pages render correctly", "pri": "P0", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": "Verify on desktop and mobile"},
        {"id": "TC-UAT-002", "cat": "UAT", "deliv": "D4", "name": "Contact form submits successfully", "desc": "Form collects name, company, email, revenue and submits", "pre": "Website live, Worker deployed", "steps": "1. Navigate to website\n2. Fill form: name, company, email, revenue\n3. Click submit\n4. Verify success message appears", "expected": "Success confirmation shown, no errors", "pri": "P0", "type": "UAT", "status": "PASS", "blocked": "", "file": "", "notes": "Form POSTs to /api/form-submit via Worker"},
        {"id": "TC-UAT-003", "cat": "UAT", "deliv": "D4", "name": "Signup page accessible at /signup", "desc": "Dedicated signup page with booking form renders", "pre": "Website live", "steps": "1. Navigate to /signup.html\n2. Verify booking form renders\n3. Check minimal navigation", "expected": "Signup page loads with booking form", "pri": "P1", "type": "UAT", "status": "PASS", "blocked": "", "file": "", "notes": "Direct link for email campaigns"},
        {"id": "TC-UAT-004", "cat": "UAT", "deliv": "D4", "name": "Website is mobile responsive", "desc": "Layout adapts to mobile screen sizes", "pre": "Website live", "steps": "1. Open website on mobile device or browser dev tools\n2. Check all 3 pages at 375px width\n3. Verify no horizontal scroll", "expected": "Layout adapts, content readable", "pri": "P1", "type": "UAT", "status": "NOT_RUN", "blocked": "", "file": "", "notes": "Test on iPhone and Android sizes"},
        {"id": "TC-UAT-005", "cat": "UAT", "deliv": "D5", "name": "Form submission creates GHL contact", "desc": "Website form submission creates contact in GoHighLevel", "pre": "GHL_API_KEY set, GHL pipeline configured", "steps": "1. Submit contact form on website\n2. Log into GoHighLevel\n3. Search for submitted email\n4. Verify contact exists with tags", "expected": "Contact found with 'cold email' tag", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "GHL_API_KEY", "file": "", "notes": "Bryce to provide GHL access"},
        {"id": "TC-UAT-006", "cat": "UAT", "deliv": "D7", "name": "Dashboard displays metric cards", "desc": "Dashboard shows email and CRM metrics", "pre": "Dashboard deployed, API connected", "steps": "1. Navigate to dashboard.html\n2. Verify Email Performance section\n3. Verify CRM Pipeline section\n4. Verify Copy Variant section", "expected": "All 3 sections render with data", "pri": "P1", "type": "UAT", "status": "PASS", "blocked": "", "file": "", "notes": "Shows demo data until APIs connected"},
        {"id": "TC-UAT-007", "cat": "UAT", "deliv": "D7", "name": "Dashboard range selector works", "desc": "Switching date ranges refreshes data", "pre": "Dashboard deployed", "steps": "1. Click 7d / 30d / All buttons\n2. Verify data values change", "expected": "Data refreshes per selected range", "pri": "P2", "type": "UAT", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-UAT-008", "cat": "UAT", "deliv": "D8", "name": "Telegram notification received for hot lead", "desc": "Hot positive reply triggers Telegram alert to team group", "pre": "TELEGRAM_BOT_TOKEN + CHAT_ID set, bot added to group", "steps": "1. Trigger hot positive reply in pipeline\n2. Check Telegram group for notification\n3. Verify message arrives within 3 hours", "expected": "Telegram message received with lead details", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "TELEGRAM_BOT_TOKEN", "file": "", "notes": "Client creates group + bot"},
        {"id": "TC-UAT-009", "cat": "UAT", "deliv": "D8", "name": "Telegram message has all lead details", "desc": "Notification includes name, email, company, industry, reply text", "pre": "Telegram notification working", "steps": "1. Trigger notification\n2. Read message\n3. Verify all fields present", "expected": "Name, email, company, industry, reply text, GHL link present", "pri": "P1", "type": "UAT", "status": "BLOCKED", "blocked": "TELEGRAM_BOT_TOKEN", "file": "", "notes": ""},
        {"id": "TC-UAT-010", "cat": "UAT", "deliv": "D6", "name": "Auto-reply sounds human (tone review)", "desc": "AI-generated replies match Aleksandar's voice", "pre": "OPENROUTER_API_KEY set, auto-reply enabled", "steps": "1. Generate 10 auto-replies via pipeline\n2. Read each reply\n3. Check: blunt, direct, no corporate speak\n4. Check: first-person 'I' voice", "expected": "All replies sound natural and match tone", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": "Bryce reviews before launch"},
        {"id": "TC-UAT-011", "cat": "UAT", "deliv": "D6", "name": "Auto-reply respects all guard rails", "desc": "No exclamation marks, <3 sentences, no AI mention", "pre": "Auto-replies generated", "steps": "1. Check 10 replies for exclamation marks\n2. Count sentences (max 3)\n3. Search for 'AI', 'automation', 'bot'", "expected": "Zero guard rail violations", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-012", "cat": "UAT", "deliv": "D6", "name": "Auto-reply delay is 2-7 minutes", "desc": "Responses arrive with realistic human-like delays", "pre": "Auto-reply system running", "steps": "1. Trigger 5 positive replies\n2. Measure time until auto-reply sent\n3. Verify each delay is 120-420 seconds", "expected": "All delays within configured range", "pri": "P1", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-013", "cat": "UAT", "deliv": "D3", "name": "AI openers reference specific business details", "desc": "Each opener mentions something specific to the prospect", "pre": "OPENROUTER_API_KEY set", "steps": "1. Generate openers for 20 diverse leads\n2. Read each opener\n3. Verify it references industry, location, or rating", "expected": "Each opener is unique and business-specific", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-014", "cat": "UAT", "deliv": "D3", "name": "AI openers match tone config", "desc": "Openers are blunt, direct, <20 words, no exclamation", "pre": "Openers generated", "steps": "1. Count words in each opener (max 20)\n2. Check for exclamation marks (zero)\n3. Verify factual observation style", "expected": "All openers meet constraints", "pri": "P1", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-015", "cat": "UAT", "deliv": "D12", "name": "Email variant copy meets constraints", "desc": "All variants are <60 words, <3 sentences, no exclamation", "pre": "email_variants.json loaded", "steps": "1. Read all variants from config\n2. Check word count, sentence count, punctuation", "expected": "All variants pass validation", "pri": "P1", "type": "UAT", "status": "PASS", "blocked": "", "file": "", "notes": "4 human variants already validated"},
        {"id": "TC-UAT-016", "cat": "UAT", "deliv": "D9", "name": "Weekly report contains accurate metrics", "desc": "Report numbers match Instantly + GHL dashboards", "pre": "INSTANTLY_API_KEY + GHL_API_KEY set", "steps": "1. Generate weekly report\n2. Compare emails_sent to Instantly dashboard\n3. Compare contacts to GHL dashboard", "expected": "Numbers match within 5% tolerance", "pri": "P1", "type": "UAT", "status": "BLOCKED", "blocked": "GHL_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-017", "cat": "UAT", "deliv": "D11", "name": "Pipeline produces reasonable lead count", "desc": "One niche + city returns sensible number of leads", "pre": "SERPER_API_KEY set", "steps": "1. Run source stage for 'car wash Houston TX'\n2. Check lead count\n3. Verify leads have valid data", "expected": "50-200 leads with business_name and domain", "pri": "P1", "type": "UAT", "status": "NOT_RUN", "blocked": "", "file": "", "notes": "SERPER_API_KEY is set -- can test now"},
        {"id": "TC-UAT-018", "cat": "UAT", "deliv": "D5", "name": "GHL pipeline stages match PRD", "desc": "Pipeline has New/Contacted/Interested/Booked/Closed stages", "pre": "GHL_API_KEY set, pipeline configured", "steps": "1. Log into GHL\n2. Navigate to pipeline settings\n3. Verify 5 stages exist", "expected": "All 5 stages present and named correctly", "pri": "P1", "type": "UAT", "status": "BLOCKED", "blocked": "GHL_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-019", "cat": "UAT", "deliv": "D6", "name": "Hot lead handoff stops AI auto-reply", "desc": "Prospect giving phone number triggers handoff, AI stops", "pre": "Auto-reply system running", "steps": "1. Send reply: 'Ready to sell, call me at 555-1234'\n2. Verify AI does NOT auto-reply\n3. Verify Telegram notification fires instead", "expected": "action=handoff, no auto-reply sent, notification fires", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-UAT-020", "cat": "UAT", "deliv": "D-ALL", "name": "Full journey walkthrough: source to notify", "desc": "End-to-end: source leads, enrich, personalize, upload, reply, classify, notify", "pre": "All API keys set", "steps": "1. Run full pipeline for 1 niche\n2. Verify leads sourced\n3. Verify emails enriched\n4. Verify openers generated\n5. Verify upload to Instantly\n6. Simulate reply\n7. Verify classification\n8. Verify GHL contact created\n9. Verify Telegram notification", "expected": "Complete flow works end-to-end", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "All API keys", "file": "", "notes": "Milestone 1 acceptance test"},

        # ── SECURITY TESTS ────────────────────────────────────────────
        {"id": "TC-SEC-001", "cat": "Security", "deliv": "D-ALL", "name": "No hardcoded API keys in source", "desc": "Scan all .py files for known key patterns", "pre": "Source files exist", "steps": "Grep execution/**/*.py for 4 forbidden key-prefix patterns (see test_sanity.py FORBIDDEN_PATTERNS)", "expected": "Zero matches", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SEC-002", "cat": "Security", "deliv": "D-ALL", "name": ".env file is gitignored", "desc": "Secrets file excluded from version control", "pre": ".gitignore exists", "steps": "Verify .env appears in .gitignore", "expected": "Found in .gitignore", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-SEC-003", "cat": "Security", "deliv": "D10", "name": "CF Worker CORS restricts origins", "desc": "Only allowed origins can access Worker endpoints", "pre": "Worker deployed", "steps": "1. Send OPTIONS request from unauthorized origin\n2. Check Access-Control-Allow-Origin header", "expected": "Unauthorized origin rejected", "pri": "P1", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": "Check wrangler.toml ALLOWED_ORIGINS"},
        {"id": "TC-SEC-004", "cat": "Security", "deliv": "D10", "name": "CF Worker form-submit validates input", "desc": "Missing required fields return error", "pre": "Worker deployed", "steps": "POST to /api/form-submit with missing name field", "expected": "Returns 400 error with message", "pri": "P1", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-SEC-005", "cat": "Security", "deliv": "D10", "name": "CF Worker rejects empty email", "desc": "Form submit with empty email rejected", "pre": "Worker deployed", "steps": "POST to /api/form-submit with empty email", "expected": "Returns 400 error", "pri": "P1", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-SEC-006", "cat": "Security", "deliv": "D4", "name": "Website form has client-side validation", "desc": "Required fields enforced before submit", "pre": "Website loaded", "steps": "1. Try submitting empty form\n2. Verify HTML5 validation fires", "expected": "Browser prevents submission, shows validation errors", "pri": "P2", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-SEC-007", "cat": "Security", "deliv": "D-ALL", "name": "No credentials in git history", "desc": "Git log does not contain secret strings", "pre": "Git repo", "steps": "git log --all -p | grep -i 'api_key\\|secret\\|password'", "expected": "No actual key values found", "pri": "P0", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": "Run before first push"},
        {"id": "TC-SEC-008", "cat": "Security", "deliv": "D10", "name": "Worker secrets stored via wrangler secret", "desc": "API keys not in wrangler.toml, use Cloudflare secrets", "pre": "Worker deployed", "steps": "Check wrangler.toml for any API key values", "expected": "No API key values in file", "pri": "P1", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-SEC-009", "cat": "Security", "deliv": "D-ALL", "name": "All API keys loaded via os.environ", "desc": "No direct key values in Python source", "pre": "Source files", "steps": "Grep for os.environ.get patterns in all scripts", "expected": "All keys use os.environ, never hardcoded", "pri": "P0", "type": "Automated", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-SEC-010", "cat": "Security", "deliv": "D6", "name": "Auto-reply never mentions AI or automation", "desc": "Guard rail prevents AI disclosure", "pre": "Auto-replies generated", "steps": "Search 50 auto-replies for 'AI', 'bot', 'automated', 'automation'", "expected": "Zero matches", "pri": "P0", "type": "UAT", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-SEC-011", "cat": "Security", "deliv": "D8", "name": "Telegram bot token not exposed in logs", "desc": "Bot token never logged or printed", "pre": "Notification system running", "steps": "Check all log outputs for token patterns", "expected": "No tokens in log files", "pri": "P1", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-SEC-012", "cat": "Security", "deliv": "D5", "name": "GHL API uses Bearer token auth", "desc": "API calls use Authorization header, not query params", "pre": "ghl.py source code", "steps": "Read ghl.py, verify Authorization header usage", "expected": "Bearer token in header, not in URL", "pri": "P1", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},

        # ── PERFORMANCE TESTS ─────────────────────────────────────────
        {"id": "TC-PERF-001", "cat": "Performance", "deliv": "D1", "name": "Serper sourcing <30s per query", "desc": "Single niche query completes in under 30 seconds", "pre": "SERPER_API_KEY set", "steps": "Time serper_maps_scraper for 1 niche", "expected": "Completes in <30 seconds", "pri": "P2", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-PERF-002", "cat": "Performance", "deliv": "D2", "name": "Email enrichment: 100 leads in <5 min", "desc": "Batch email lookup completes in reasonable time", "pre": "ANYMAILFINDER_API_KEY set", "steps": "Run enrichment on 100 leads, measure time", "expected": "Completes in <5 minutes", "pri": "P2", "type": "Manual", "status": "BLOCKED", "blocked": "ANYMAILFINDER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-PERF-003", "cat": "Performance", "deliv": "D3", "name": "Opener generation: 50 leads in <3 min", "desc": "Batch opener generation completes quickly", "pre": "OPENROUTER_API_KEY set", "steps": "Run opener generator on 50 leads, measure time", "expected": "Completes in <3 minutes", "pri": "P2", "type": "Manual", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-PERF-004", "cat": "Performance", "deliv": "D11", "name": "Full mock pipeline <60s", "desc": "Complete pipeline run in mock mode finishes quickly", "pre": "None", "steps": "Time py execution/gtm_client_workflows/accessory_masters_pipeline.py --mock --force", "expected": "Completes in <60 seconds", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "", "notes": "Currently completes in ~2s"},
        {"id": "TC-PERF-005", "cat": "Performance", "deliv": "D-ALL", "name": "Deduplication: 1000 leads in <1s", "desc": "Dedup function scales to expected volume", "pre": "None", "steps": "Generate 1000 mock leads, time deduplicate()", "expected": "Completes in <1 second", "pri": "P2", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-PERF-006", "cat": "Performance", "deliv": "D10", "name": "CF Worker response <500ms", "desc": "API endpoints respond quickly", "pre": "Worker deployed", "steps": "curl -w '%{time_total}' each endpoint", "expected": "All responses <500ms", "pri": "P1", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-PERF-007", "cat": "Performance", "deliv": "D6", "name": "Reply classification <2s per reply", "desc": "Single reply classification is fast", "pre": "OPENROUTER_API_KEY set", "steps": "Time classify() on 10 replies", "expected": "Average <2 seconds per reply", "pri": "P2", "type": "Manual", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-PERF-008", "cat": "Performance", "deliv": "D11", "name": "Pipeline handles 800 leads without OOM", "desc": "Daily target volume doesn't exhaust memory", "pre": "None", "steps": "Run pipeline with 800 mock leads", "expected": "No memory errors, completes normally", "pri": "P1", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-PERF-009", "cat": "Performance", "deliv": "D12", "name": "Variant generation <5s", "desc": "Single AI variant generated quickly", "pre": "OPENROUTER_API_KEY set", "steps": "Time generate_challenger_variant()", "expected": "Completes in <5 seconds", "pri": "P3", "type": "Manual", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-PERF-010", "cat": "Performance", "deliv": "D9", "name": "Report generation <10s", "desc": "Weekly report generates quickly", "pre": "API keys set", "steps": "Time run_weekly_report(mock=True)", "expected": "Completes in <10 seconds", "pri": "P3", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": ""},

        # ── DEPLOYMENT VALIDATION ─────────────────────────────────────
        {"id": "TC-DEPL-001", "cat": "Deployment", "deliv": "D4", "name": "CF Pages returns 200 OK", "desc": "Website serves without errors", "pre": "CF Pages deployed", "steps": "curl -s -o /dev/null -w '%{http_code}' https://accessory-masters-site.pages.dev", "expected": "Returns 200", "pri": "P0", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-002", "cat": "Deployment", "deliv": "D10", "name": "CF Worker health check", "desc": "Worker base URL responds", "pre": "Worker deployed", "steps": "curl https://accessory-masters-api.accessory-masters.workers.dev/", "expected": "Returns valid response", "pri": "P0", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-003", "cat": "Deployment", "deliv": "D10", "name": "/api/form-submit endpoint reachable", "desc": "Form endpoint accepts POST requests", "pre": "Worker deployed", "steps": "POST to /api/form-submit with test data", "expected": "Returns 200 or expected response", "pri": "P0", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-004", "cat": "Deployment", "deliv": "D10", "name": "/api/dashboard endpoint reachable", "desc": "Dashboard data endpoint responds", "pre": "Worker deployed", "steps": "GET /api/dashboard", "expected": "Returns JSON with metric data", "pri": "P1", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-005", "cat": "Deployment", "deliv": "D10", "name": "/api/webhook/reply endpoint reachable", "desc": "Reply webhook endpoint accepts POST", "pre": "Worker deployed", "steps": "POST to /api/webhook/reply with test payload", "expected": "Returns 200", "pri": "P1", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-006", "cat": "Deployment", "deliv": "D10", "name": "/api/variants endpoint reachable", "desc": "Variant data endpoint responds", "pre": "Worker deployed", "steps": "GET /api/variants", "expected": "Returns JSON", "pri": "P2", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-007", "cat": "Deployment", "deliv": "D4", "name": "_redirects proxies /api/* to Worker", "desc": "API calls from Pages routed to Worker", "pre": "CF Pages + Worker deployed", "steps": "curl https://accessory-masters-site.pages.dev/api/dashboard", "expected": "Returns Worker response (not 404)", "pri": "P1", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-008", "cat": "Deployment", "deliv": "D10", "name": "wrangler.toml correct CF account", "desc": "Worker config points to right account", "pre": "wrangler.toml exists", "steps": "Read file, verify account_id matches", "expected": "account_id = 26e5b8612be35e5d23a9186fcf5288d0", "pri": "P1", "type": "Automated", "status": "PASS", "blocked": "", "file": "tests/test_sanity.py", "notes": ""},
        {"id": "TC-DEPL-009", "cat": "Deployment", "deliv": "D4", "name": "All 3 HTML pages serve without errors", "desc": "index, dashboard, signup all return 200", "pre": "CF Pages deployed", "steps": "Fetch each page, verify HTTP 200", "expected": "All 3 return 200", "pri": "P0", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},
        {"id": "TC-DEPL-010", "cat": "Deployment", "deliv": "D-ALL", "name": "Python dependencies install cleanly", "desc": "pip install works without errors", "pre": "Python 3.14 installed", "steps": "pip install -r requirements.txt (or manual installs)", "expected": "All packages install successfully", "pri": "P1", "type": "Manual", "status": "PASS", "blocked": "", "file": "", "notes": ""},

        # ── LIVE API INTEGRATION TESTS ────────────────────────────────
        {"id": "TC-LIVE-001", "cat": "Live API", "deliv": "D1", "name": "Serper live: car wash Houston TX", "desc": "Real Serper API returns Houston car wash businesses", "pre": "SERPER_API_KEY set", "steps": "Run serper_maps_scraper.py for 'car wash Houston TX'", "expected": "Returns 20+ leads with business_name, address, domain", "pri": "P0", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": "Key is SET -- can test now"},
        {"id": "TC-LIVE-002", "cat": "Live API", "deliv": "D1", "name": "Prospeo live: domain contact search", "desc": "Real Prospeo API returns contacts for a domain", "pre": "PROSPEO_API_KEY set", "steps": "Call search-person endpoint with a known domain", "expected": "Returns contact data or empty (Houston coverage is limited)", "pri": "P2", "type": "Manual", "status": "NOT_RUN", "blocked": "", "file": "", "notes": "Key is SET -- can test now"},
        {"id": "TC-LIVE-003", "cat": "Live API", "deliv": "D2", "name": "AnymailFinder live: domain email lookup", "desc": "Real AnymailFinder returns email for a business domain", "pre": "ANYMAILFINDER_API_KEY set", "steps": "Call AnymailFinder API with a known business domain", "expected": "Returns email with confidence score", "pri": "P0", "type": "Manual", "status": "BLOCKED", "blocked": "ANYMAILFINDER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-LIVE-004", "cat": "Live API", "deliv": "D2", "name": "Million Verifier live: email verification", "desc": "Real Million Verifier validates an email address", "pre": "MILLION_VERIFIER_API_KEY set + credits > 0", "steps": "Call Million Verifier API with a known email", "expected": "Returns verification status (ok/invalid/catch_all)", "pri": "P0", "type": "Manual", "status": "BLOCKED", "blocked": "Million Verifier credits (0 balance)", "file": "", "notes": "Key SET but 0 credits"},
        {"id": "TC-LIVE-005", "cat": "Live API", "deliv": "D3", "name": "AI opener live: generates personalized opener", "desc": "Real LLM generates business-specific opening line", "pre": "OPENROUTER_API_KEY set", "steps": "Call generate_opener with real lead data", "expected": "Returns personalized, <20 word opener", "pri": "P0", "type": "Manual", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
        {"id": "TC-LIVE-006", "cat": "Live API", "deliv": "D5", "name": "GHL live: create test contact", "desc": "Real GHL API creates a contact in the CRM", "pre": "GHL_API_KEY set", "steps": "POST to GHL /contacts/ with test data", "expected": "Contact created, ID returned", "pri": "P0", "type": "Manual", "status": "BLOCKED", "blocked": "GHL_API_KEY", "file": "", "notes": ""},
        {"id": "TC-LIVE-007", "cat": "Live API", "deliv": "D8", "name": "Telegram live: send test notification", "desc": "Real Telegram Bot API sends message to group", "pre": "TELEGRAM_BOT_TOKEN + CHAT_ID set", "steps": "Call notify_positive_reply with test data", "expected": "Message appears in Telegram group", "pri": "P0", "type": "Manual", "status": "BLOCKED", "blocked": "TELEGRAM_BOT_TOKEN", "file": "", "notes": ""},
        {"id": "TC-LIVE-008", "cat": "Live API", "deliv": "D6", "name": "Reply classifier live: LLM classifies correctly", "desc": "Real LLM classifies sample replies accurately", "pre": "OPENROUTER_API_KEY set", "steps": "Call classify() with 7 sample replies (hot, positive, negative, neutral)", "expected": "90%+ accuracy (6/7 correct)", "pri": "P0", "type": "Manual", "status": "BLOCKED", "blocked": "OPENROUTER_API_KEY", "file": "", "notes": ""},
    ]


# ---------------------------------------------------------------------------
# Deliverable metadata for traceability matrix
# ---------------------------------------------------------------------------

DELIVERABLES = [
    ("D1", "Lead Sourcing (Serper Maps + Prospeo)"),
    ("D2", "Email Enrichment & Verification (AnymailFinder + Million Verifier)"),
    ("D3", "AI Personalization (LLM Openers)"),
    ("D4", "Website (CF Pages: index, dashboard, signup)"),
    ("D5", "CRM Integration (GoHighLevel)"),
    ("D6", "Reply Classification & Auto-Reply System"),
    ("D7", "Dashboard (client-facing metrics)"),
    ("D8", "Telegram Notifications"),
    ("D9", "Weekly Report Generator"),
    ("D10", "Cloudflare Worker (API Proxy)"),
    ("D11", "Pipeline Orchestration (end-to-end)"),
    ("D12", "Self-Optimizing Copy Loop (Variant Generator)"),
    ("D-ALL", "Cross-Cutting / System-Wide"),
    ("D-INFRA", "Infrastructure & DevOps"),
]


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1a1a2e")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1a1a2e")
LABEL_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)

PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "P1": PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
    "P2": PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid"),
    "P3": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
}

STATUS_FILLS = {
    "PASS": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
    "FAIL": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
    "BLOCKED": PatternFill(start_color="FFF0D0", end_color="FFF0D0", fill_type="solid"),
    "NOT_RUN": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _apply_header_row(ws, row: int, headers: list[str], widths: list[int] | None = None):
    """Write a header row with dark styling and optional column widths."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def _apply_body_cell(ws, row: int, col: int, value, wrap: bool = True):
    """Write a body cell with standard formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP_ALIGNMENT if wrap else CENTER_ALIGNMENT
    return cell


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_executive_summary(wb: Workbook, test_cases: list[dict]):
    """Sheet 1: Executive Summary."""
    ws = wb.active
    ws.title = "Executive Summary"

    # Title row (rows 1-2 merged)
    ws.merge_cells("A1:N2")
    title_cell = ws.cell(row=1, column=1, value="Accessory Masters — QA Test Plan")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Project info (rows 4-9)
    info = [
        ("Project", "Cold Email System for Business Acquisition Outreach"),
        ("Client", "Aleksandar & Simon (Accessory Masters)"),
        ("Manager", "Bryce Lindberg (DoubleClick AI)"),
        ("Builder", "Debanjan Mazumdar (AntiGravity)"),
        ("Date", "May 1, 2026"),
        ("Target", "May 11, 2026"),
    ]
    for i, (label, value) in enumerate(info):
        row = 4 + i
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = BODY_FONT
        val_cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)

    # Test Coverage Summary (row 11+)
    ws.cell(row=11, column=1, value="Test Coverage Summary").font = SUBTITLE_FONT

    summary_headers = ["Category", "Total", "Automated", "Manual/UAT", "Pass", "Blocked"]
    _apply_header_row(ws, 12, summary_headers, widths=[24, 10, 14, 14, 10, 10])

    # Build category summary
    cat_counter: dict[str, dict] = {}
    for tc in test_cases:
        cat = tc["cat"]
        if cat not in cat_counter:
            cat_counter[cat] = {"total": 0, "automated": 0, "manual_uat": 0, "pass": 0, "blocked": 0}
        cat_counter[cat]["total"] += 1
        if tc["type"] == "Automated":
            cat_counter[cat]["automated"] += 1
        else:
            cat_counter[cat]["manual_uat"] += 1
        if tc["status"] == "PASS":
            cat_counter[cat]["pass"] += 1
        if tc["status"] == "BLOCKED":
            cat_counter[cat]["blocked"] += 1

    # Stable category order
    cat_order = [
        "Unit Test", "Integration", "E2E Pipeline", "Sanity",
        "Edge Case", "UAT", "Security", "Performance",
        "Deployment", "Live API",
    ]
    row = 13
    totals = {"total": 0, "automated": 0, "manual_uat": 0, "pass": 0, "blocked": 0}
    for cat in cat_order:
        if cat not in cat_counter:
            continue
        d = cat_counter[cat]
        vals = [cat, d["total"], d["automated"], d["manual_uat"], d["pass"], d["blocked"]]
        for col_idx, v in enumerate(vals, 1):
            _apply_body_cell(ws, row, col_idx, v, wrap=False)
        for k in totals:
            totals[k] += d[k]
        row += 1

    # Totals row
    total_vals = ["TOTAL", totals["total"], totals["automated"], totals["manual_uat"], totals["pass"], totals["blocked"]]
    for col_idx, v in enumerate(total_vals, 1):
        cell = _apply_body_cell(ws, row, col_idx, v, wrap=False)
        cell.font = LABEL_FONT
    row += 2

    # Priority Definitions
    ws.cell(row=row, column=1, value="Priority Definitions").font = SECTION_FONT
    row += 1
    pri_defs = [
        ("P0", "Blocker — system cannot launch without this"),
        ("P1", "Critical — major feature gap, needs fix before launch"),
        ("P2", "Important — should fix, but launch not blocked"),
        ("P3", "Nice-to-have — cosmetic or minor improvement"),
    ]
    for code, desc in pri_defs:
        ws.cell(row=row, column=1, value=code).font = LABEL_FONT
        ws.cell(row=row, column=2, value=desc).font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
    row += 1

    # Status Definitions
    ws.cell(row=row, column=1, value="Status Definitions").font = SECTION_FONT
    row += 1
    status_defs = [
        ("PASS", "Test executed and succeeded"),
        ("FAIL", "Test executed and failed — defect logged"),
        ("BLOCKED", "Cannot execute — dependency missing (see Blocked By column)"),
        ("NOT_RUN", "Not yet executed — scheduled for upcoming cycle"),
        ("N/A", "Not applicable to current configuration"),
    ]
    for code, desc in status_defs:
        ws.cell(row=row, column=1, value=code).font = LABEL_FONT
        ws.cell(row=row, column=2, value=desc).font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    # Column widths for readability
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60


def build_test_cases_sheet(wb: Workbook, test_cases: list[dict]):
    """Sheet 2: Test Cases (main sheet)."""
    ws = wb.create_sheet("Test Cases")

    headers = [
        "Test ID", "Category", "Deliverable", "Test Name", "Description",
        "Preconditions", "Steps", "Expected Result", "Priority", "Type",
        "Current Status", "Blocked By", "Automated Test File", "Notes",
    ]
    widths = [14, 20, 12, 40, 50, 30, 50, 40, 8, 12, 14, 22, 30, 30]

    _apply_header_row(ws, 1, headers, widths)

    # Data validation for Priority, Type, Status
    dv_priority = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
    dv_priority.error = "Please select P0, P1, P2, or P3"
    dv_priority.errorTitle = "Invalid Priority"
    ws.add_data_validation(dv_priority)

    dv_type = DataValidation(type="list", formula1='"Automated,Manual,UAT"', allow_blank=True)
    dv_type.error = "Please select Automated, Manual, or UAT"
    dv_type.errorTitle = "Invalid Type"
    ws.add_data_validation(dv_type)

    dv_status = DataValidation(type="list", formula1='"PASS,FAIL,BLOCKED,NOT_RUN,N/A"', allow_blank=True)
    dv_status.error = "Please select a valid status"
    dv_status.errorTitle = "Invalid Status"
    ws.add_data_validation(dv_status)

    # Column key mapping
    keys = ["id", "cat", "deliv", "name", "desc", "pre", "steps", "expected",
            "pri", "type", "status", "blocked", "file", "notes"]

    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, key in enumerate(keys, 1):
            cell = _apply_body_cell(ws, row_idx, col_idx, tc.get(key, ""))
            # Priority color
            if key == "pri" and tc.get("pri") in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[tc["pri"]]
                cell.alignment = CENTER_ALIGNMENT
            # Status color
            if key == "status" and tc.get("status") in STATUS_FILLS:
                cell.fill = STATUS_FILLS[tc["status"]]
                cell.alignment = CENTER_ALIGNMENT

        # Apply data validation to the specific cells in this row
        dv_priority.add(ws.cell(row=row_idx, column=9))
        dv_type.add(ws.cell(row=row_idx, column=10))
        dv_status.add(ws.cell(row=row_idx, column=11))

    # Auto-filter
    last_row = len(test_cases) + 1
    ws.auto_filter.ref = f"A1:N{last_row}"

    # Freeze header row
    ws.freeze_panes = "A2"


def build_traceability_matrix(wb: Workbook, test_cases: list[dict]):
    """Sheet 3: Traceability Matrix."""
    ws = wb.create_sheet("Traceability Matrix")

    headers = ["Deliverable", "Description", "Test IDs", "Test Count", "Coverage Notes"]
    widths = [14, 55, 80, 12, 40]
    _apply_header_row(ws, 1, headers, widths)

    # Group test IDs by deliverable
    deliv_map: dict[str, list[str]] = defaultdict(list)
    for tc in test_cases:
        deliv_map[tc["deliv"]].append(tc["id"])

    row = 2
    for deliv_code, deliv_desc in DELIVERABLES:
        ids = deliv_map.get(deliv_code, [])
        count = len(ids)
        coverage = "Covered" if count > 0 else "NO TESTS - needs coverage"
        _apply_body_cell(ws, row, 1, deliv_code, wrap=False)
        _apply_body_cell(ws, row, 2, deliv_desc)
        _apply_body_cell(ws, row, 3, ", ".join(ids))
        _apply_body_cell(ws, row, 4, count, wrap=False)
        note_cell = _apply_body_cell(ws, row, 5, coverage)
        if count == 0:
            note_cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        row += 1

    ws.freeze_panes = "A2"


def build_blocked_items(wb: Workbook, test_cases: list[dict]):
    """Sheet 4: Blocked Items."""
    ws = wb.create_sheet("Blocked Items")

    headers = ["Test ID", "Test Name", "Blocked By", "Who Provides",
               "Expected Date", "Impact", "Status"]
    widths = [14, 40, 28, 20, 16, 40, 12]
    _apply_header_row(ws, 1, headers, widths)

    # Map common blockers to who provides them
    provider_map = {
        "GHL_API_KEY": "Bryce / Client",
        "TELEGRAM_BOT_TOKEN": "Client (Aleksandar)",
        "OPENROUTER_API_KEY": "Debanjan",
        "ANYMAILFINDER_API_KEY": "Debanjan",
        "Million Verifier credits (0 balance)": "Debanjan (top-up)",
        "All API keys": "Bryce + Client + Debanjan",
    }

    blocked = [tc for tc in test_cases if tc.get("blocked")]
    row = 2
    for tc in blocked:
        blocker = tc["blocked"]
        provider = provider_map.get(blocker, "TBD")
        _apply_body_cell(ws, row, 1, tc["id"], wrap=False)
        _apply_body_cell(ws, row, 2, tc["name"])
        _apply_body_cell(ws, row, 3, blocker)
        _apply_body_cell(ws, row, 4, provider)
        _apply_body_cell(ws, row, 5, "TBD", wrap=False)
        _apply_body_cell(ws, row, 6, f"Blocks {tc['cat']} testing for {tc['deliv']}")
        _apply_body_cell(ws, row, 7, "BLOCKED", wrap=False)
        ws.cell(row=row, column=7).fill = STATUS_FILLS["BLOCKED"]
        row += 1

    ws.freeze_panes = "A2"


def build_execution_log(wb: Workbook):
    """Sheet 5: Execution Log (empty template)."""
    ws = wb.create_sheet("Execution Log")

    headers = ["Test ID", "Tester", "Date", "Status", "Actual Result",
               "Screenshots/Evidence", "Defects Found", "Resolution"]
    widths = [14, 18, 14, 12, 50, 30, 30, 30]
    _apply_header_row(ws, 1, headers, widths)

    # 50 empty rows with light borders
    for row in range(2, 52):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate Accessory Masters QA Test Plan (XLSX)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=".tmp/accessory_masters_qa_test_plan.xlsx",
        help="Output path for the XLSX file (default: .tmp/accessory_masters_qa_test_plan.xlsx)",
    )
    args = parser.parse_args()

    output_path = ROOT / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    test_cases = build_all_test_cases()

    wb = Workbook()
    build_executive_summary(wb, test_cases)
    build_test_cases_sheet(wb, test_cases)
    build_traceability_matrix(wb, test_cases)
    build_blocked_items(wb, test_cases)
    build_execution_log(wb)

    wb.save(str(output_path))

    # Summary
    status_counts = Counter(tc["status"] for tc in test_cases)
    type_counts = Counter(tc["type"] for tc in test_cases)
    blocked_count = sum(1 for tc in test_cases if tc.get("blocked"))

    print(f"QA Test Plan generated: {output_path}")
    print(f"  Total test cases: {len(test_cases)}")
    print(f"  PASS: {status_counts.get('PASS', 0)}  |  FAIL: {status_counts.get('FAIL', 0)}  |  BLOCKED: {status_counts.get('BLOCKED', 0)}  |  NOT_RUN: {status_counts.get('NOT_RUN', 0)}")
    print(f"  Automated: {type_counts.get('Automated', 0)}  |  Manual: {type_counts.get('Manual', 0)}  |  UAT: {type_counts.get('UAT', 0)}")
    print(f"  Blocked items (needing action): {blocked_count}")


if __name__ == "__main__":
    main()
